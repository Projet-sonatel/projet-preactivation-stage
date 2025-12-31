import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import datetime

# Configuration de la page
st.set_page_config(page_title="Classement PVT ", layout="wide")

# Titre
st.title("📊 Classement des PVT - 7 Directions Régionales")

# Liste des 7 DR autorisées
DR_AUTORISEES = [
    "DV-DRVN_DIRECTION REGIONALE DES VENTES NORD",
    "DV-DRVC_DIRECTION REGIONALE DES VENTES CENTRE",
    "DV-DRV1_DIRECTION REGIONALE DES VENTES DAKAR 1",
    "DV-DRV2_DIRECTION REGIONALE DES VENTES DAKAR 2",
    "DV-DRVS_DIRECTION REGIONALE DES VENTES SUD",
    "DV-DRVSE_DIRECTION REGIONALE DES VENTES SUD-EST",
    "DV-DRVE_DIRECTION REGIONALE DES VENTES EST"
]

# Mapping pour les codes DR courts
DR_MAPPING = {
    "DV-DRVN_DIRECTION REGIONALE DES VENTES NORD": "DRN",
    "DV-DRVC_DIRECTION REGIONALE DES VENTES CENTRE": "DRC",
    "DV-DRV1_DIRECTION REGIONALE DES VENTES DAKAR 1": "DR1",
    "DV-DRV2_DIRECTION REGIONALE DES VENTES DAKAR 2": "DR2",
    "DV-DRVS_DIRECTION REGIONALE DES VENTES SUD": "DRS",
    "DV-DRVSE_DIRECTION REGIONALE DES VENTES SUD-EST": "DRSE",
    "DV-DRVE_DIRECTION REGIONALE DES VENTES EST": "DRE"
}

# Fonctions
def filter_pvt(df):
    df_filtered = df.copy()
    df_filtered['PVT'] = df_filtered['PVT'].astype(str).str.strip()
    mask = df_filtered['PVT'].str.upper().str.startswith('PVT')
    return df_filtered[mask]

def filter_etat_identification(df):
    df_filtered = df.copy()
    if 'ETAT_IDENTIFICATION' not in df_filtered.columns:
        return df_filtered
    df_filtered['ETAT_IDENTIFICATION'] = df_filtered['ETAT_IDENTIFICATION'].astype(str).str.strip()
    mask = df_filtered['ETAT_IDENTIFICATION'].str.contains("Identifie Photo", case=False, na=False)
    return df_filtered[mask]

def get_telephone_by_pvt(df):
    if 'MSISDN' not in df.columns:
        return pd.Series([None] * len(df))
    df['MSISDN_CLEAN'] = df['MSISDN'].astype(str).str.strip()
    df['MSISDN_CLEAN'] = df['MSISDN_CLEAN'].str.replace(r'\.0$', '', regex=True)
    telephone_series = df.groupby('PVT')['MSISDN_CLEAN'].first()
    return df['PVT'].map(telephone_series)

def generate_excel_classement(df_classement):
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_classement.to_excel(writer, sheet_name='Classement PVT', index=False)

    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active

    # Styles
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True, size=11, color="000000")
    header_alignment = Alignment(horizontal='center', vertical='center')

    data_font = Font(size=10)
    data_alignment_left = Alignment(horizontal='left', vertical='center')
    data_alignment_center = Alignment(horizontal='center', vertical='center')

    total_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
    total_font = Font(bold=True, size=11, color="000000")

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Appliquer les styles
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(min_row=2, max_row=max_row-1, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            cell.font = data_font
            if cell.column == 1 or cell.column == 9:
                cell.alignment = data_alignment_center
            else:
                cell.alignment = data_alignment_left

    total_row = max_row
    for cell in ws[total_row]:
        cell.fill = total_fill
        cell.font = total_font
        cell.border = thin_border
        if cell.column in [1, 9]:
            cell.alignment = data_alignment_center

    # Largeurs de colonnes
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 12

    ws.freeze_panes = 'A2'

    final_buffer = BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)
    return final_buffer

# Interface
uploaded_file = st.file_uploader("", type=["xlsx", "csv"])

if uploaded_file:
    try:
        if uploaded_file.name.endswith('.csv'):
            try:
                df = pd.read_csv(uploaded_file, sep='|', encoding='utf-8')
            except:
                try:
                    df = pd.read_csv(uploaded_file, sep=';', encoding='utf-8')
                except:
                    df = pd.read_csv(uploaded_file, sep=',', encoding='utf-8')
        else:
            df = pd.read_excel(uploaded_file)

        # Mapping des colonnes
        column_mapping = {
            'ACCUEIL_VENDEUR': 'PVT',
            'AGENCE_VENDEUR': 'DR',
            'LOGIN_VENDEUR': 'LOGIN',
            'MSISDN': 'MSISDN',
            'ETAT_IDENTIFICATION': 'ETAT_IDENTIFICATION',
            'PRENOM_VENDEUR': 'PRENOM_VENDEUR',
            'NOM_VENDEUR': 'NOM_VENDEUR'
        }

        for old_name, new_name in column_mapping.items():
            if old_name in df.columns and new_name not in df.columns:
                df = df.rename(columns={old_name: new_name})

        if 'PRENOM_VENDEUR' not in df.columns:
            df['PRENOM_VENDEUR'] = ''
        if 'NOM_VENDEUR' not in df.columns:
            df['NOM_VENDEUR'] = ''

        required_columns = ['PVT', 'DR', 'LOGIN', 'MSISDN']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            st.error(f"❌ Colonnes manquantes : {', '.join(missing_columns)}")
        else:
            with st.spinner("⏳ Traitement en cours..."):
                # Filtrage
                df_filtre_dr = df[df['DR'].isin(DR_AUTORISEES)].copy()
                if len(df_filtre_dr) == 0:
                    st.error("❌ Aucune donnée trouvée pour les 7 DR spécifiées.")
                    st.stop()

                df_filtre_etat = filter_etat_identification(df_filtre_dr)
                df_filtre_pvt = filter_pvt(df_filtre_etat)

                if len(df_filtre_pvt) == 0:
                    st.error("❌ Aucun PVT ne commence par 'PVT' dans les données filtrées.")
                    st.stop()

                # Nettoyage
                df_filtre_pvt['PVT'] = df_filtre_pvt['PVT'].astype(str).str.strip()
                df_filtre_pvt['DR'] = df_filtre_pvt['DR'].astype(str).str.strip()
                df_filtre_pvt['LOGIN'] = df_filtre_pvt['LOGIN'].astype(str).str.strip()
                df_filtre_pvt['MSISDN'] = df_filtre_pvt['MSISDN'].astype(str).str.strip()
                df_filtre_pvt['PRENOM_VENDEUR'] = df_filtre_pvt['PRENOM_VENDEUR'].astype(str).str.strip()
                df_filtre_pvt['NOM_VENDEUR'] = df_filtre_pvt['NOM_VENDEUR'].astype(str).str.strip()

                if 'ETAT_IDENTIFICATION' in df_filtre_pvt.columns:
                    df_filtre_pvt['ETAT_IDENTIFICATION'] = df_filtre_pvt['ETAT_IDENTIFICATION'].astype(str).str.strip()

                # Téléphone
                df_filtre_pvt['TELEPHONE'] = get_telephone_by_pvt(df_filtre_pvt)

                # Groupement
                group_cols = ['DR', 'PVT', 'LOGIN', 'PRENOM_VENDEUR', 'NOM_VENDEUR', 'TELEPHONE']
                if 'ETAT_IDENTIFICATION' in df_filtre_pvt.columns:
                    group_cols.append('ETAT_IDENTIFICATION')

                df_grouped = df_filtre_pvt.groupby(group_cols).size().reset_index(name='VENTES_TOTALES')

                # Codes DR courts
                df_grouped['DR_COURT'] = df_grouped['DR'].map(DR_MAPPING)
                df_grouped['DR'] = df_grouped['DR_COURT'].fillna(df_grouped['DR'])
                df_grouped = df_grouped.drop(columns=['DR_COURT'])

                # Tri et classement
                df_grouped = df_grouped.sort_values('VENTES_TOTALES', ascending=False)
                df_grouped['RANG'] = range(1, len(df_grouped) + 1)

                # Organisation des colonnes
                columns_order = ['RANG', 'PVT', 'LOGIN', 'PRENOM_VENDEUR', 'NOM_VENDEUR', 'DR', 'TELEPHONE']
                if 'ETAT_IDENTIFICATION' in df_grouped.columns:
                    columns_order.append('ETAT_IDENTIFICATION')
                columns_order.append('VENTES_TOTALES')

                df_classement = df_grouped[columns_order]

                # Total
                total_ventes = df_classement['VENTES_TOTALES'].sum()
                df_display = df_classement.copy()
                total_row = ['', 'TOTAL', '', '', '', '', '']
                if 'ETAT_IDENTIFICATION' in df_classement.columns:
                    total_row.append('')
                total_row.append(total_ventes)
                df_display.loc[len(df_display)] = total_row

                # Génération Excel
                excel_file = generate_excel_classement(df_classement)
                date_str = datetime.now().strftime("%Y%m%d_%H%M")
                filename = f"Classement_PVT{date_str}.xlsx"

                # Téléchargement
                st.download_button(
                    label="📥 Télécharger le fichier Excel",
                    data=excel_file,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    except Exception as e:
        st.error(f"❌ Erreur : {str(e)}")